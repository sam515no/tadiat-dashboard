"""
generate.py — يقرأ ملف الإكسل ويولد index.html للداشبورد
يُشغَّل تلقائياً عبر GitHub Actions عند رفع إكسل جديد
"""

import openpyxl, json, math, sys, os, glob

# ── إيجاد ملف الإكسل ─────────────────────────────────────────────
xlsx_files = glob.glob("data/*.xlsx")
if not xlsx_files:
    print("❌ لا يوجد ملف إكسل في مجلد data/")
    sys.exit(1)
XLSX_PATH = xlsx_files[0]
print(f"📂 تحميل: {XLSX_PATH}")

# ── قراءة الإكسل ─────────────────────────────────────────────────
wb = openpyxl.load_workbook(XLSX_PATH)
ws = wb.active

def parse_coord(v):
    if v is None or v == '_': return None
    try: return float(v)
    except:
        try: return float(str(v).replace(' ', ''))
        except: return None

data = []
for r in range(4, ws.max_row + 1):
    num = ws.cell(r, 1).value
    if num is None: continue
    d_obs = ws.cell(r, 4).value
    date_str = d_obs.strftime("%Y-%m-%d") if hasattr(d_obs, 'strftime') else str(d_obs or '')
    E = parse_coord(ws.cell(r, 8).value)
    N = parse_coord(ws.cell(r, 9).value)
    vtype = str(ws.cell(r, 13).value or '').strip()
    if 'سطح' in vtype or 'تخفيض' in vtype: vtype = 'سطحي'
    elif 'امطار' in vtype or 'أمطار' in vtype: vtype = 'امطار'
    elif 'قناة' in vtype or 'قنوات' in vtype: vtype = 'قناة'

    data.append({
        "num": int(num),
        "ref": str(ws.cell(r, 2).value or ''),
        "desc": str(ws.cell(r, 3).value or ''),
        "date_obs": date_str,
        "E": E, "N": N,
        "municipality": str(ws.cell(r, 10).value or '').strip(),
        "neighborhood": str(ws.cell(r, 11).value or '').strip(),
        "street": str(ws.cell(r, 12).value or '').strip(),
        "type": vtype or 'غير محدد',
        "status": str(ws.cell(r, 16).value or 'قائم').strip(),
        "balance": ""
    })

print(f"✅ {len(data)} تعدي")

# ── تصنيف التعديات حسب القرب من الشبكات ─────────────────────────
with open("template.html", encoding="utf-8") as f:
    tmpl = f.read()

# Extract NETWORK_LINES from template
import re
m = re.search(r'const NETWORK_LINES = (\[.*?\]);', tmpl, re.DOTALL)
net = json.loads(m.group(1)) if m else []
surface_lines = [n["coords"] for n in net if n["type"] == "سطحي"]
rain_lines    = [n["coords"] for n in net if n["type"] == "امطار"]

def pt_seg_dist(px, py, ax, ay, bx, by):
    dx, dy = bx-ax, by-ay
    if dx == 0 and dy == 0: return math.hypot(px-ax, py-ay)
    t = max(0, min(1, ((px-ax)*dx + (py-ay)*dy) / (dx*dx + dy*dy)))
    return math.hypot(px-(ax+t*dx), py-(ay+t*dy))

def min_dist(lng, lat, lines):
    best = float('inf')
    for coords in lines:
        for i in range(len(coords)-1):
            d = pt_seg_dist(lng, lat, coords[i][0], coords[i][1], coords[i+1][0], coords[i+1][1])
            if d < best: best = d
    return best

THRESH = 0.003
classification = {}
for d in data:
    if not d["E"] or not d["N"]: continue
    ds = min_dist(d["E"], d["N"], surface_lines)
    dr = min_dist(d["E"], d["N"], rain_lines)
    if ds < THRESH or dr < THRESH:
        classification[str(d["num"])] = {"classification": "سطحي" if ds <= dr else "امطار"}
    d["type"] = classification.get(str(d["num"]), {}).get("classification", d["type"])

from collections import Counter
types = Counter(d["type"] for d in data)
print(f"   سطحي: {types.get('سطحي',0)} | امطار: {types.get('امطار',0)}")

# ── حقن البيانات في القالب ────────────────────────────────────────
total       = len(data)
with_coords = sum(1 for d in data if d["E"] and d["N"])
surface_cnt = types.get("سطحي", 0)
rain_cnt    = types.get("امطار", 0)

html = tmpl
html = html.replace("DATA_PLACEHOLDER",    json.dumps(data, ensure_ascii=False))
html = html.replace("CLASSIF_PLACEHOLDER", json.dumps(classification, ensure_ascii=False))

# Update stat counters
html = re.sub(r'id="totalCount">\d+<',   f'id="totalCount">{total}<',       html)
html = re.sub(r'id="activeCount">\d+<',  f'id="activeCount">{total}<',      html)
html = re.sub(r'id="surfaceCount">\d+<', f'id="surfaceCount">{surface_cnt}<', html)
html = re.sub(r'id="rainCount">\d+<',    f'id="rainCount">{rain_cnt}<',      html)
html = re.sub(r'>\d+</div><div class="sl">بإحداثيات', f'>{with_coords}</div><div class="sl">بإحداثيات', html)

with open("index.html", "w", encoding="utf-8") as f:
    f.write(html)

print(f"✅ index.html جاهز ({len(html)//1024} KB)")
